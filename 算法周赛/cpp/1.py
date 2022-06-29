# coding=utf-8

# import openpyxl

from openpyxl import load_workbook
wb = load_workbook('/Users/mac/Documents/2021_my_media/project/algorithm-leetcode/算法周赛/cpp/OD.xlsx')
ws = wb['OD']

resStr = ''

# 1814
for i in range(2, 1814):
    tempList = []
    tempStr = "{"
    for j in range(1, 7):
        # 注： str()
        val = str(ws.cell(i, j).value)
        tempVal = r'"'
        tempVal += (val + '"')
        tempList.append(tempVal)
        
        # if (i > 1810):
        # print(tempVal)
        
        # {"1", "万州", "上海", "25.8", "1500", "0.13"},

    tempStr += (', '.join(tempList))
    tempStr += "}"
    resStr += (tempStr + ",\n")


    # if i > 1812:
    #     print(tempStr)
    # print(resStr)
    # break

print(resStr)

print(r'"sss"')
